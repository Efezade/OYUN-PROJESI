# -*- coding: utf-8 -*-
"""MD dosyasindaki tum tablolari bir .xlsx workbook'una (her tablo ayri sheet) aktarir.
Sheet adi: tabloya en yakin ustteki baslik (#, ##, ###) metninden turetilir.
Kullanim: python md_tables_to_xlsx.py <kaynak.md> [<hedef.xlsx>]
Hedef verilmezse kaynakla ayni klasorde, ayni adla .xlsx uretilir.
"""
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

if len(sys.argv) < 2:
    print("Kullanim: python md_tables_to_xlsx.py <kaynak.md> [<hedef.xlsx>]")
    sys.exit(1)

SRC = Path(sys.argv[1]).resolve()
OUT = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else SRC.with_suffix(".xlsx")

with open(SRC, encoding="utf-8") as f:
    lines = f.read().split("\n")

n = len(lines)
i = 0
current_heading = "Sheet"
tables = []  # list of (heading, header_row, body_rows)
used_names = set()


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]\:\*\?/\\]", " ", name).strip()
    name = re.sub(r"\s+", " ", name)
    if not name:
        name = "Sheet"
    base = name[:31]
    candidate = base
    k = 2
    while candidate in used_names:
        suffix = f" ({k})"
        candidate = base[: 31 - len(suffix)] + suffix
        k += 1
    used_names.add(candidate)
    return candidate


while i < n:
    line = lines[i]
    stripped = line.strip()

    m = re.match(r"^(#{1,3})\s+(.*)$", stripped)
    if m:
        current_heading = m.group(2).strip()
        i += 1
        continue

    if stripped.startswith("|") and i + 1 < n and re.match(r"^\|[\s:\-|]+\|?\s*$", lines[i + 1].strip()):
        tbl_rows = []
        while i < n and lines[i].strip().startswith("|"):
            tbl_rows.append(lines[i].strip())
            i += 1
        header = [c.strip() for c in tbl_rows[0].strip("|").split("|")]
        body = []
        for row in tbl_rows[2:]:
            cells = [c.strip() for c in row.strip("|").split("|")]
            body.append(cells)
        tables.append((current_heading, header, body))
        continue

    i += 1

if not tables:
    print("UYARI: tablo bulunamadi, xlsx uretilmedi ->", SRC)
    sys.exit(0)

wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for heading, header, body in tables:
    ws = wb.create_sheet(sanitize_sheet_name(heading))
    ws.append(header)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in body:
        row = row + [""] * (len(header) - len(row))
        ws.append(row[: len(header)])
    for col_idx, col_name in enumerate(header, start=1):
        max_len = max([len(str(col_name))] + [len(str(r[col_idx - 1])) for r in body if col_idx - 1 < len(r)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
    ws.freeze_panes = "A2"

wb.save(OUT)
print("OK ->", OUT, f"({len(tables)} tablo)")
